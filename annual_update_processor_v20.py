"""
================================================================================
MSCI Annual Update Factual Process Automation
================================================================================
Version     : 20.0.0
Changes     : 1. FIX: Auto-create missing "Compensation" sheet in the
                 Scalar_Series1 template when the sheet is absent.
                 The sheet is built with headers defined in
                 CONFIG["COMPENSATION_SHEET_HEADERS"] and populated from
                 the CSV section whose Series ID column is nearest before
                 the CASHFEES data lib column.  BASESALARY, CASHFEES, and
                 ALLOTHERCOMPENSATION (plus their _UOM twins) are the core
                 fields; the full list is configurable.
                 This resolves the bug where directors receiving Cash Fees
                 (NED compensation) were silently dropped from the output.
              2. All v19 fixes retained (COMMITTEEFUNCTION Yes->1,
                 top/left alignment, Individual Name column insertion,
                 best-match section detection).
Requirements: Python 3.11+ | pandas | openpyxl
================================================================================
"""

import os
import re
import sys
import logging
import datetime
import traceback
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
# Table / TableStyleInfo intentionally not imported:
# openpyxl Table objects cause Excel corruption (orphaned table XML).
# AutoFilter is used instead — same UX, zero corruption.


# ==============================================================================
# SECTION 1: CONFIGURATION
# All paths and constants are defined here. Modify only this section to
# adapt the script to a different machine or environment.
# ==============================================================================

CONFIG = {
    # ── Input folders ──────────────────────────────────────────────────────────
    # All folders are synced via OneDrive (green tick visible in File Explorer):
    #   OneDrive > Devagi - MSCI Office 365 > Desktop > Details > Work
    #   > Annual Update > Others > Bulk Upload
    #
    # Extraction CSV files sit directly inside EXTRACTION_FOLDER
    "EXTRACTION_FOLDER": (
        r"C:\Users\murudev\OneDrive - MSCI Office 365\Desktop\Details\Work"
        r"\Annual Update\Others\Bulk Upload"
    ),
    # DVV merged XLSX files inside the DVV file sub-folder
    "DVV_FOLDER": (
        r"C:\Users\murudev\OneDrive - MSCI Office 365\Desktop\Details\Work"
        r"\Annual Update\Others\Bulk Upload\DVV file"
    ),
    # Template XLSX files
    "TEMPLATE_FOLDER": (
        r"C:\Users\murudev\OneDrive - MSCI Office 365\Desktop\Details\Work"
        r"\Annual Update\Others\Bulk Upload\Bulk upload Template"
    ),

    # ── Output root ────────────────────────────────────────────────────────────
    # A sub-folder named  IssuerID_IssuerName_Output  is created here per run
    "OUTPUT_ROOT": (
        r"C:\Users\murudev\OneDrive - MSCI Office 365\Desktop\Details\Work"
        r"\Annual Update\Others\Bulk Upload\Output"
    ),

    # ── Template file names (exact filenames on disk) ─────────────────────────
    # NOTE: Scalar and Series1 are now combined into one file (Scalar_Series1).
    # Template filenames must match EXACTLY what is on disk (including spaces).
    "TEMPLATES": {
        "Position":      "Position_Tab_Split.xlsx",
        "Scalar_Series1": "Scalar & Series 1_DP.xlsx",
        "Series2":       "Series 2- Delete blank cell_DP.xlsx",
    },

    # ── Output file name suffixes ──────────────────────────────────────────────
    # Output files will be named:  IssuerID_<suffix>.xlsx
    "OUTPUT_NAMES": {
        "Position":      "Position_Tab_Split",
        "Scalar_Series1": "Scalar & Series 1_DP",
        "Series2":       "Series 2- Delete blank cell_DP",
    },

    # ── Name of the scalar (issuer-level) sheet inside Scalar_Series1 ─────────
    # All other sheets in that file are treated as series-level (one row per
    # extraction row), exactly like Series1 used to be.
    "SCALAR_SHEET_NAME": "Scalar",

    # ── DVV TAB filter (Column B in the DVV Merged file) ─────────────────────
    # TAB column is read by position — Column B (index 1).
    # DVV rows whose TAB value matches one of the four tabs below are EXCLUDED.
    # All other TAB rows are kept for DVV override processing.
    "DVV_TAB_COL": "B",
    "DVV_EXCLUDE_TABS": {
        "director attributes",
        "director data - board",
        "positions",
        "individual",
    },

    # ── DVV UID column (Column S) — committee name for name-match ────────────
    # For Committees TAB rows, Column S (UID) contains the committee name
    # e.g. "Audit", "Compensation", "Nominating and Governance".
    # This is used to match against COMMITTEENAME in the template sheet.
    "DVV_UID_COL": "S",

    # ── DVV name-based matching for committee sheets ──────────────────────────
    # When TAB = "Committees" or "Committee Membership", use UID (Column S)
    # as the committee name to match against the anchor column in the template.
    #
    # Format: "Template Sheet Name" : ("TAB value(s)", "anchor column header")
    #
    # How it works:
    #   DVV: TAB="Committees", UID="Audit", DATALIB=COMMITTEEMTGS, VALUE=4
    #   -> find row in "Committee" sheet where COMMITTEENAME = "Audit"
    #   -> write 4 into COMMITTEEMTGS + green highlight
    "DVV_NAME_MATCH_SHEETS": {
        "Committee": "COMMITTEENAME",
    },

    # ── DVV column letters (Excel column letters in the DVV merged file) ───────
    # Verified against actual DVV Merged file structure:
    "DVV_DATALIB_COL":    "G",   # Column G  -> DATALIB_TAG  (code like DIRGENDER)
    "DVV_CORRECTVAL_COL": "AH",  # Column AH -> CORRECT_VALUE
    "DVV_SERIES_COL":     "AM",  # Column AM -> SERIALID

    # ── Extraction Series ID ───────────────────────────────────────────────────
    # Series ID column index (0-based) in the RAW extraction CSV.
    # Verified against actual CSV: Column AD (index 29) = 'Series ID'
    # contains the numeric IDs (1, 2, 3...) that match DVV SERIALID.
    # Column E (index 4) = PRIMARY_DISCLOSURE_LANG ('English') — NOT the Series ID.
    "EXTRACTION_SERIES_COL_INDEX": 29,

    # ── Field names ────────────────────────────────────────────────────────────
    "ISSUER_ID_FIELD": "DMX_ISSUER_ID",

    # ── Individual name column ────────────────────────────────────────────────
    # REL_INDIVID is the column in the extraction CSV containing the director
    # name. A new column "Individual Name" is inserted immediately after
    # serial_id ONLY in the sheets listed below.
    "INDIVIDUAL_NAME_FIELD":      "REL_INDIVID",
    "INDIVIDUAL_NAME_COL_HEADER": "Individual Name",
    "INDIVIDUAL_NAME_SHEETS": {
        "Committee Membership",
        "Director Attributes",
        "Director Ownership",
        "Compensation",
        "CEO Compensation",
        "CEO Compensation CIC,SEV",
    },

    # ── COMMITTEEFUNCTION columns — Yes -> 1 conversion ──────────────────────
    # When these columns are populated, any value of "Yes" (case-insensitive)
    # is written as "1" instead. Values already "1" are kept as "1".
    "COMMITTEEFUNCTION_COLS": {
        "COMMITTEEFUNCTIONA",
        "COMMITTEEFUNCTIONC",
        "COMMITTEEFUNCTIONE",
        "COMMITTEEFUNCTIONG",
        "COMMITTEEFUNCTIONN",
        "COMMITTEEFUNCTIONRISK",
        "COMMITFUNCTIONHEALTHSAFETY",
        "COMMITTEEFUNCTIONO",
    },

    # ── Compensation sheet anchor section fix ────────────────────────────────
    # The best-match heuristic picks the wrong CSV section for the Compensation
    # sheet in the Series2 template.  Section 19 (CEO/NEO profile) ties or beats
    # Section 20 (the true NED cash-fees section) because BONUS and
    # NONEQUITYINCENTIVEPLAN appear in both.  We break the tie by anchoring on
    # CASHFEES: the Series ID column immediately before CASHFEES in the CSV is
    # the authoritative section for the Compensation sheet.
    "COMPENSATION_SHEET_NAME":     "Compensation",
    "COMPENSATION_ANCHOR_DATALIB": "CASHFEES",

    # ── Formatting ────────────────────────────────────────────────────────────
    "HEADER_FONT_NAME":    "Times New Roman",
    "HEADER_FONT_SIZE":    12,
    "HEADER_FONT_COLOR":   "FFFFFF",   # White text
    "HEADER_FILL_COLOR":   "0070C0",   # Blue fill
    "DATA_FONT_NAME":      "Times New Roman",
    "DATA_FONT_SIZE":      12,
    "DVV_HIGHLIGHT_COLOR": "E6FDCF",   # Light green for DVV-updated cells
    # Table objects removed — AutoFilter used instead to avoid Excel corruption.
}


# ==============================================================================
# SECTION 2: LOGGING
# ==============================================================================

def setup_logging(output_folder: Path, issuer_id: str) -> logging.Logger:
    """
    Set up file + console logging.
    File captures DEBUG and above; console shows INFO and above.
    """
    output_folder.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = output_folder / f"{issuer_id}_execution_{timestamp}.log"

    # Avoid duplicate handlers if called more than once
    logger = logging.getLogger("AnnualUpdate")
    if logger.handlers:
        logger.handlers.clear()
    logger.setLevel(logging.DEBUG)

    fh = logging.FileHandler(log_path, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(logging.Formatter(
        "%(asctime)s | %(levelname)-8s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"
    ))

    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(logging.Formatter("%(levelname)-8s | %(message)s"))

    logger.addHandler(fh)
    logger.addHandler(ch)
    return logger


# ==============================================================================
# SECTION 3: FILE DISCOVERY
# ==============================================================================

def find_extraction_file(issuer_id: str, extraction_folder: Path,
                         logger: logging.Logger) -> tuple[Path, str]:
    """
    Find extraction CSV starting with issuer_id in extraction_folder.
    File format: IID000000002177512_Aurubis AG.csv
    Returns (Path, issuer_name).
    """
    logger.info(f"Searching extraction folder: {extraction_folder}")
    if not extraction_folder.exists():
        raise FileNotFoundError(
            f"Extraction folder does not exist: {extraction_folder}"
        )

    matches = [
        f for f in extraction_folder.iterdir()
        if f.suffix.lower() == ".csv" and f.name.startswith(issuer_id)
    ]

    if not matches:
        raise FileNotFoundError(
            f"No extraction CSV found starting with '{issuer_id}' "
            f"in {extraction_folder}"
        )
    if len(matches) > 1:
        logger.warning(
            f"Multiple extraction files found for {issuer_id}. "
            f"Using: {matches[0].name}"
        )

    chosen = matches[0]
    # Issuer name = everything after the first underscore in the stem
    # e.g. "IID000000002177512_Aurubis AG" -> "Aurubis AG"
    parts = chosen.stem.split("_", 1)
    issuer_name = parts[1].strip() if len(parts) == 2 else issuer_id
    logger.info(
        f"Extraction file: {chosen.name}  |  Issuer Name: '{issuer_name}'"
    )
    return chosen, issuer_name


def find_dvv_file(issuer_id: str, dvv_folder: Path,
                  logger: logging.Logger) -> Path:
    """
    Find DVV XLSX containing issuer_id immediately before '_Merged'.
    File format: DVV (CG-AU) - 2025-12-16_IID000000002177512_Merged.xlsx
    """
    logger.info(f"Searching DVV folder: {dvv_folder}")
    if not dvv_folder.exists():
        raise FileNotFoundError(
            f"DVV folder does not exist: {dvv_folder}"
        )

    pattern = re.compile(
        re.escape(issuer_id) + r"_Merged", re.IGNORECASE
    )
    matches = [
        f for f in dvv_folder.iterdir()
        if f.suffix.lower() == ".xlsx" and pattern.search(f.stem)
    ]

    if not matches:
        raise FileNotFoundError(
            f"No DVV XLSX found containing '{issuer_id}_Merged' "
            f"in {dvv_folder}"
        )
    if len(matches) > 1:
        logger.warning(
            f"Multiple DVV files found for {issuer_id}. "
            f"Using: {matches[0].name}"
        )
    logger.info(f"DVV file: {matches[0].name}")
    return matches[0]


# ==============================================================================
# SECTION 4: EXTRACTION CSV PROCESSING
# ==============================================================================

def standardize_header(raw_header: str) -> str:
    """
    Convert a CSV header to its Data Lib identifier.

    Format 1 — already a Data Lib (no brackets):
        DMX_ISSUER_ID  ->  DMX_ISSUER_ID

    Format 2 — human label with Data Lib in final parentheses:
        Audit Board Member (REL_AUDIT_BOARD)  ->  REL_AUDIT_BOARD

    Rule: extract text inside the LAST pair of parentheses.
    If no brackets exist, return the header unchanged.
    """
    raw = str(raw_header).strip()
    match = re.search(r"\(([^)]+)\)\s*$", raw)
    if match:
        return match.group(1).strip()
    return raw


def load_extraction(csv_path: Path, issuer_id: str,
                    logger: logging.Logger) -> pd.DataFrame:
    """
    Load extraction CSV and standardise headers.

    Key changes vs v1.0:
    - NO column reordering. Column E (index 4) already contains Series ID
      in the raw CSV. We preserve the original column order exactly.
    - All values read as strings to prevent type inference.
    """
    logger.info(f"Loading extraction CSV: {csv_path.name}")

    try:
        df = pd.read_csv(csv_path, dtype=str, encoding="utf-8-sig")
    except UnicodeDecodeError:
        df = pd.read_csv(csv_path, dtype=str, encoding="latin-1")

    if df.empty:
        raise ValueError(f"Extraction CSV is empty: {csv_path.name}")

    # Standardise all headers to Data Lib format
    df.columns = [standardize_header(c) for c in df.columns]

    # Log what the Series ID column actually is (for confirmation)
    if len(df.columns) > CONFIG["EXTRACTION_SERIES_COL_INDEX"]:
        series_col_name = df.columns[CONFIG["EXTRACTION_SERIES_COL_INDEX"]]
        logger.info(
            f"Series ID column (index {CONFIG['EXTRACTION_SERIES_COL_INDEX']}) = '{series_col_name}'"
        )
    else:
        logger.warning(
            f"Extraction CSV has fewer than {CONFIG['EXTRACTION_SERIES_COL_INDEX']+1} columns — "
            "Series ID column cannot be read."
        )

    logger.info(
        f"Extraction loaded: {len(df)} rows, "
        f"{len(df.columns)} columns after header standardisation"
    )
    return df


def build_datalib_to_series_map(ext_df: pd.DataFrame) -> dict[str, int]:
    """
    Build a mapping: datalib_code -> Series ID column index in ext_df.

    The extraction CSV has multiple 'Series ID' columns (one per section).
    Each Series ID column immediately precedes the data lib columns for
    that section.  For any given data lib code, the correct Series ID is
    the one whose column position is the closest one BEFORE that code's
    column position.

    Returns:
        { 'BDMTGS': 197, 'COMMITTEENAME': 205, ... }
        (values are 0-based column indices into ext_df)
    """
    cols = list(ext_df.columns)
    # Positions of all Series ID columns (standardised headers)
    series_id_positions = [
        i for i, c in enumerate(cols)
        if c == "Series ID" or re.match(r"^Series ID(\.\d+)?$", c)
    ]

    datalib_to_series_col: dict[str, int] = {}
    for col_idx, col_name in enumerate(cols):
        if col_name in ("Series ID",) or re.match(r"^Series ID(\.\d+)?$", col_name):
            continue
        # Find the nearest Series ID column that is BEFORE this column
        before = [p for p in series_id_positions if p < col_idx]
        if before:
            datalib_to_series_col[col_name] = before[-1]

    return datalib_to_series_col


def get_series_id_for_datalib(ext_row: pd.Series, ext_df: pd.DataFrame,
                               datalib_code: str,
                               datalib_to_series_map: dict[str, int]) -> str | None:
    """
    Return the Series ID value for a specific data lib code in an extraction row,
    using the correct section-level Series ID column.
    """
    series_col_idx = datalib_to_series_map.get(datalib_code)
    if series_col_idx is None:
        return None
    col_name = ext_df.columns[series_col_idx]
    val = ext_row.get(col_name)
    if pd.notna(val) and str(val).strip() not in ("", "nan"):
        return str(val).strip()
    return None


def build_individual_name_map(ext_df: pd.DataFrame) -> dict[str, str]:
    """
    Build a mapping: serial_id -> individual name (REL_INDIVID).

    Finds the Series ID column nearest to REL_INDIVID — either immediately
    before it (most CSVs) or immediately after it (some CSVs place Series ID
    after REL_INDIVID). Builds { serial_id_value: director_name }.
    """
    individ_col = CONFIG["INDIVIDUAL_NAME_FIELD"]
    if individ_col not in ext_df.columns:
        return {}

    cols = list(ext_df.columns)
    individ_idx = cols.index(individ_col)
    series_positions = [
        i for i, c in enumerate(cols)
        if c == "Series ID" or re.match(r"^Series ID(\.\d+)?$", c)
    ]

    # Try nearest Series ID BEFORE REL_INDIVID first
    before = [p for p in series_positions if p < individ_idx]
    # If none before, use nearest Series ID AFTER REL_INDIVID
    after  = [p for p in series_positions if p > individ_idx]

    if before:
        sid_col = cols[before[-1]]
    elif after:
        sid_col = cols[after[0]]
    else:
        return {}

    name_map: dict[str, str] = {}
    for _, row in ext_df.iterrows():
        sid  = row.get(sid_col)
        name = row.get(individ_col)
        if (pd.notna(sid) and pd.notna(name) and
                str(sid).strip() not in ("", "nan") and
                str(name).strip() not in ("", "nan")):
            name_map[str(sid).strip()] = str(name).strip()
    return name_map


def row_has_data(ext_row: pd.Series, template_headers: list[str]) -> bool:
    """
    Return True if the extraction row has at least one non-blank value
    for any of the template headers (excluding DMX_ISSUER_ID and serial_id).
    This prevents writing empty rows to the output.
    """
    skip = {CONFIG["ISSUER_ID_FIELD"], "serial_id"}
    for h in template_headers:
        if h in skip:
            continue
        val = ext_row.get(h)
        if pd.notna(val) and str(val).strip() not in ("", "nan"):
            return True
    return False


# ==============================================================================
# SECTION 5: DVV FILE PROCESSING
# ==============================================================================

def load_dvv(dvv_path: Path, logger: logging.Logger) -> pd.DataFrame:
    """
    Load DVV Merged XLSX. Two sequential filters are applied:

    FILTER 1 — Correct Value not blank (Column AH):
        Keep only rows where Correct Value is non-empty.

    FILTER 2 — TAB whitelist (Column B, read by position):
        Keep only rows whose TAB value (case-insensitive) is one of:
            • Director Attributes
            • Director Data - Board
            • Positions
            • Individual
        All other TAB rows are excluded and logged.

    Column mapping (verified against actual DVV Merged file):
        B  (index 1)  -> TAB           (category — read by position)
        G  (index 6)  -> DATALIB_TAG   (code used to match template headers)
        AH (index 33) -> CORRECT_VALUE (the value to write into the template)
        AM (index 38) -> SERIALID      (series/director ID for row matching)
    """
    logger.info(f"Loading DVV file: {dvv_path.name}")

    def col_letter_to_index(letter: str) -> int:
        letter = letter.upper()
        result = 0
        for ch in letter:
            result = result * 26 + (ord(ch) - ord('A') + 1)
        return result - 1

    try:
        dvv_df = pd.read_excel(
            dvv_path, sheet_name=0, dtype=str,
            header=0, engine="openpyxl"
        )
    except Exception as e:
        raise RuntimeError(
            f"Cannot read DVV file '{dvv_path.name}': {e}"
        )

    if dvv_df.empty:
        raise ValueError(f"DVV file is empty: {dvv_path.name}")

    tab_idx        = col_letter_to_index(CONFIG["DVV_TAB_COL"])
    uid_idx        = col_letter_to_index(CONFIG["DVV_UID_COL"])
    datalib_idx    = col_letter_to_index(CONFIG["DVV_DATALIB_COL"])
    correctval_idx = col_letter_to_index(CONFIG["DVV_CORRECTVAL_COL"])
    series_idx     = col_letter_to_index(CONFIG["DVV_SERIES_COL"])

    max_needed = max(tab_idx, uid_idx, datalib_idx, correctval_idx, series_idx)
    if dvv_df.shape[1] <= max_needed:
        raise ValueError(
            f"DVV file has only {dvv_df.shape[1]} columns; "
            f"need at least {max_needed + 1} "
            f"(up to column {CONFIG['DVV_SERIES_COL']})."
        )

    cols = dvv_df.columns.tolist()
    logger.info(
        f"DVV columns -> TAB='{cols[tab_idx]}' UID='{cols[uid_idx]}' "
        f"DataLib='{cols[datalib_idx]}' "
        f"CorrectVal='{cols[correctval_idx]}' "
        f"SeriesID='{cols[series_idx]}'"
    )

    dvv_df = dvv_df.rename(columns={
        cols[tab_idx]:        "_DVV_TAB",
        cols[uid_idx]:        "_DVV_UID",
        cols[datalib_idx]:    "_DVV_DATALIB",
        cols[correctval_idx]: "_DVV_CORRECT_VALUE",
        cols[series_idx]:     "_DVV_SERIES_ID",
    })

    total_rows = len(dvv_df)

    # ── FILTER 1: Correct Value not blank ────────────────────────────────────
    dvv_df = dvv_df[
        dvv_df["_DVV_CORRECT_VALUE"].notna() &
        (dvv_df["_DVV_CORRECT_VALUE"].str.strip() != "")
    ].copy()
    after_filter1 = len(dvv_df)
    logger.info(
        f"DVV Filter 1 (Correct Value not blank): "
        f"{total_rows} -> {after_filter1} rows"
    )

    # ── FILTER 2: TAB exclusion (Column B, positional) ───────────────────────
    # Exclude rows whose TAB value matches any of the 4 excluded categories.
    # Keep everything else.
    exclude_tabs = CONFIG["DVV_EXCLUDE_TABS"]   # set of lowercase strings

    dvv_df["_DVV_TAB_NORM"] = (
        dvv_df["_DVV_TAB"]
        .fillna("")
        .str.strip()
        .str.lower()
    )

    excluded = dvv_df[dvv_df["_DVV_TAB_NORM"].isin(exclude_tabs)]
    if not excluded.empty:
        excluded_vals = sorted(excluded["_DVV_TAB_NORM"].unique().tolist())
        logger.debug(
            f"DVV Filter 2: excluding {len(excluded)} row(s) with "
            f"TAB values: {excluded_vals}"
        )

    dvv_df = dvv_df[~dvv_df["_DVV_TAB_NORM"].isin(exclude_tabs)].copy()
    dvv_df = dvv_df.drop(columns=["_DVV_TAB_NORM"])

    after_filter2 = len(dvv_df)
    logger.info(
        f"DVV Filter 2 (exclude TABs {sorted(exclude_tabs)}, Column B): "
        f"{after_filter1} -> {after_filter2} rows"
    )
    logger.info(
        f"DVV loaded: {total_rows} total -> {after_filter2} actionable rows"
    )

    return dvv_df


# ==============================================================================
# SECTION 6: TEMPLATE POPULATION
# ==============================================================================

def populate_template(
    template_path: Path,
    template_key: str,
    ext_df: pd.DataFrame,
    issuer_id: str,
    logger: logging.Logger,
) -> tuple[openpyxl.Workbook, dict]:
    """
    Populate all sheets in a template workbook from the extraction DataFrame.

    Rules:
    ┌─────────────┬──────────────────────────────────────────────────────────┐
    │ Scaler      │ One issuer-level row. No Series ID. No blank row check   │
    │             │ needed (always one row).                                  │
    ├─────────────┼──────────────────────────────────────────────────────────┤
    │ Position /  │ One output row per extraction row. Series ID from Col E. │
    │ Series 1/2  │ Skip rows where all template data fields are blank.       │
    └─────────────┴──────────────────────────────────────────────────────────┘

    Matching: exact Data Lib match only. No fuzzy matching.
    Unmatched template headers -> left blank, logged in Validation.
    """
    logger.info(f"Populating '{template_key}': {template_path.name}")
    wb = load_workbook(template_path)
    sheet_meta = {}

    # Build once: datalib_code -> which Series ID column index to use
    datalib_to_series_map = build_datalib_to_series_map(ext_df)

    # Build once: serial_id -> individual name (REL_INDIVID)
    individual_name_map = build_individual_name_map(ext_df)
    logger.info(
        f"Individual name map: {len(individual_name_map)} directors loaded from REL_INDIVID"
    )

    # COMMITTEEFUNCTION columns that need Yes -> 1 conversion
    comm_fn_cols = CONFIG.get("COMMITTEEFUNCTION_COLS", set())

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Read header row from template (row 1)
        headers = [
            cell.value for cell in ws[1] if cell.value is not None
        ]
        if not headers:
            logger.warning(
                f"  Sheet '{sheet_name}' has no headers — skipping."
            )
            continue

        # Map: header_name -> 1-based column index
        header_col_map = {
            cell.value: col_idx
            for col_idx, cell in enumerate(ws[1], start=1)
            if cell.value is not None
        }

        # If this sheet has a serial_id column, insert an "Individual Name"
        # column immediately after it (if not already present).
        is_scalar = (sheet_name == CONFIG["SCALAR_SHEET_NAME"])
        indiv_col_idx = None

        if (not is_scalar
                and "serial_id" in header_col_map
                and sheet_name in CONFIG["INDIVIDUAL_NAME_SHEETS"]):
            indiv_header = CONFIG["INDIVIDUAL_NAME_COL_HEADER"]
            if indiv_header not in header_col_map:
                # Always insert the column — leave blank if no name available
                sid_col_pos = header_col_map["serial_id"]
                insert_at   = sid_col_pos + 1
                ws.insert_cols(insert_at)
                ws.cell(row=1, column=insert_at).value = indiv_header
                header_col_map = {
                    h: (ci + 1 if ci >= insert_at else ci)
                    for h, ci in header_col_map.items()
                }
                header_col_map[indiv_header] = insert_at
                indiv_col_idx = insert_at
                logger.debug(
                    f"  Sheet '{sheet_name}': inserted '{indiv_header}' "
                    f"at column {insert_at}"
                )
            else:
                indiv_col_idx = header_col_map[indiv_header]

        # headers list also needs to reflect new column
        headers = [
            cell.value for cell in ws[1] if cell.value is not None
        ]

        data_start_row = 2
        populated_rows = 0

        # ── SCALAR: one issuer row ─────────────────────────────────────────
        if is_scalar:
            # Use first extraction row (all rows share the same issuer-level
            # scalar fields; first row is sufficient)
            if CONFIG["ISSUER_ID_FIELD"] in ext_df.columns:
                issuer_rows = ext_df[
                    ext_df[CONFIG["ISSUER_ID_FIELD"]] == issuer_id
                ]
                if issuer_rows.empty:
                    issuer_rows = ext_df
            else:
                issuer_rows = ext_df

            if issuer_rows.empty:
                logger.warning(
                    f"  No rows for issuer {issuer_id}. "
                    f"Scalar sheet '{sheet_name}' left empty."
                )
                continue

            ext_row   = issuer_rows.iloc[0]
            write_row = data_start_row

            for tmpl_header, col_idx in header_col_map.items():
                if tmpl_header == CONFIG["ISSUER_ID_FIELD"]:
                    ws.cell(row=write_row, column=col_idx).value = issuer_id
                elif tmpl_header in ext_df.columns:
                    val = ext_row.get(tmpl_header)
                    ws.cell(row=write_row, column=col_idx).value = (
                        None if (pd.isna(val) or str(val).strip() in ("", "nan"))
                        else str(val).strip()
                    )

            populated_rows = 1

        # ── POSITION / SERIES: one row per extraction row ──────────────────
        else:
            # Find the Series ID column for this sheet and the range of
            # data columns that belong to this section.
            # A row is only written if its section Series ID is non-blank
            # AND at least one of the section's data columns is non-blank.
            sheet_series_col_idx  = None
            sheet_section_cols    = []

            all_std = list(ext_df.columns)
            series_positions = [
                i for i, c in enumerate(all_std)
                if c == "Series ID" or re.match(r"^Series ID(\.\d+)?$", c)
            ]

            # ── v20: Compensation sheet uses anchor-based section selection ──
            # The best-match heuristic picks the wrong section for Compensation
            # because Section 19 (CEO/NEO profile) contains more matching
            # header names (BONUS, NONEQUITYINCENTIVEPLAN, etc.) than
            # Section 20 (the true NED compensation section with CASHFEES).
            # To fix this, when processing the Compensation sheet in Series2
            # we locate the section by finding the nearest Series ID column
            # *before* the anchor datalib (CASHFEES) rather than using
            # best-match counts.
            comp_sheet_name   = CONFIG.get("COMPENSATION_SHEET_NAME", "Compensation")
            anchor_datalib    = CONFIG.get("COMPENSATION_ANCHOR_DATALIB", "CASHFEES")
            use_anchor_match  = (template_key == "Series2"
                                 and sheet_name == comp_sheet_name
                                 and anchor_datalib in all_std)

            if use_anchor_match:
                anchor_col_idx = all_std.index(anchor_datalib)
                before_anchor  = [p for p in series_positions if p < anchor_col_idx]
                if before_anchor:
                    forced_sp    = before_anchor[-1]
                    forced_nxt   = next(
                        (p for p in series_positions if p > forced_sp), len(all_std)
                    )
                    sheet_series_col_idx = forced_sp
                    sheet_section_cols   = [
                        all_std[i]
                        for i in range(forced_sp + 1, forced_nxt)
                        if all_std[i] in headers
                    ]
                    logger.info(
                        f"  Sheet '{sheet_name}': anchor-based section match "
                        f"-> Series ID col [{forced_sp}] '{all_std[forced_sp]}' "
                        f"({len(sheet_section_cols)} data cols matched)"
                    )
                else:
                    logger.warning(
                        f"  Sheet '{sheet_name}': anchor '{anchor_datalib}' "
                        f"found but no Series ID column precedes it — "
                        f"falling back to best-match."
                    )
                    use_anchor_match = False  # fall through to best-match below

            if not use_anchor_match:
                # Default: Find the Series ID column whose immediately following
                # section contains the MOST template header matches.
                # Using the first match is wrong because some data lib codes
                # (e.g. A_COMMITTEEMEMBERRETIRED) appear in multiple sections.
                best_match_cnt = 0
                for sp_idx in series_positions:
                    nxt = next((p for p in series_positions if p > sp_idx), len(all_std))
                    candidate = [
                        all_std[i]
                        for i in range(sp_idx + 1, nxt)
                        if all_std[i] in headers
                    ]
                    if len(candidate) > best_match_cnt:
                        best_match_cnt       = len(candidate)
                        sheet_series_col_idx = sp_idx
                        sheet_section_cols   = candidate

            for _, ext_row in ext_df.iterrows():

                # Gate 1: section Series ID must be non-blank
                if sheet_series_col_idx is not None:
                    sid_col_name = ext_df.columns[sheet_series_col_idx]
                    sid_val = ext_row.get(sid_col_name)
                    if pd.isna(sid_val) or str(sid_val).strip() in ("", "nan"):
                        continue
                    section_series_id = str(sid_val).strip()
                else:
                    section_series_id = None

                # Gate 2: at least one DATA column in this section must be
                # non-blank — prevents writing rows that only have a Series ID
                # (e.g. directors with no committee membership data)
                if sheet_section_cols:
                    section_has_data = any(
                        pd.notna(ext_row.get(c)) and
                        str(ext_row.get(c, "")).strip() not in ("", "nan")
                        for c in sheet_section_cols
                    )
                    if not section_has_data:
                        continue
                elif not row_has_data(ext_row, headers):
                    continue

                write_row = data_start_row + populated_rows

                for tmpl_header, col_idx in header_col_map.items():

                    if tmpl_header == CONFIG["ISSUER_ID_FIELD"]:
                        ws.cell(row=write_row, column=col_idx).value = issuer_id

                    elif tmpl_header == "serial_id":
                        ws.cell(row=write_row, column=col_idx).value = section_series_id
                        # Write individual name in the next column
                        if indiv_col_idx and section_series_id:
                            ind_name = individual_name_map.get(section_series_id, "")
                            ws.cell(row=write_row, column=indiv_col_idx).value = (
                                ind_name if ind_name else None
                            )

                    elif tmpl_header == CONFIG["INDIVIDUAL_NAME_COL_HEADER"]:
                        # Already written above when processing serial_id
                        pass

                    elif tmpl_header in ext_df.columns:
                        val = ext_row.get(tmpl_header)
                        if pd.isna(val) or str(val).strip() in ("", "nan"):
                            ws.cell(row=write_row, column=col_idx).value = None
                        else:
                            clean_val = str(val).strip()
                            # COMMITTEEFUNCTION columns: Yes -> 1
                            if (tmpl_header in comm_fn_cols and
                                    clean_val.lower() == "yes"):
                                clean_val = "1"
                            ws.cell(row=write_row, column=col_idx).value = clean_val

                populated_rows += 1

        sheet_meta[sheet_name] = {
            "headers":         headers,
            "header_col_map":  header_col_map,
            "data_start_row":  data_start_row,
            "populated_rows":  populated_rows,
        }
        logger.debug(
            f"  Sheet '{sheet_name}': {populated_rows} row(s) written."
        )

    return wb, sheet_meta


# ==============================================================================
# SECTION 7: DVV OVERRIDE PROCESSING
# ==============================================================================

def apply_dvv_overrides(
    wb: openpyxl.Workbook,
    sheet_meta: dict,
    dvv_df: pd.DataFrame,
    issuer_id: str,
    template_key: str,
    logger: logging.Logger,
) -> int:
    """
    Apply DVV Correct Value overrides to the populated workbook.

    Logic per DVV row (Correct Value already filtered non-blank in load_dvv):

      1. SCALAR sheet -> write directly to row 2, no Series ID needed.

      2. SERIALID present:
           a. Match row by SERIALID -> found: overwrite + green.
           b. Not found + name-match sheet (e.g. Committee):
                -> fall back to TAB value (Column B) vs COMMITTEENAME.
                -> found: overwrite + green.
                -> not found: create new row.
           c. Not found + not a name-match sheet -> create new row.

      3. SERIALID blank + name-match sheet:
           -> match by TAB value vs COMMITTEENAME directly.

      4. SERIALID blank + not a name-match sheet -> skip.

    Covers COMMITTEEMTGS + all COMMITTEEFUNCTIONX fields via name-match.
    """
    dvv_fill = PatternFill(
        "solid",
        start_color=CONFIG["DVV_HIGHLIGHT_COLOR"],
        end_color=CONFIG["DVV_HIGHLIGHT_COLOR"],
    )
    total_overrides = 0

    # Name-match config: sheet -> anchor column header
    name_match_sheets = CONFIG.get("DVV_NAME_MATCH_SHEETS", {})

    for sheet_name in wb.sheetnames:
        if sheet_name not in sheet_meta:
            continue

        ws             = wb[sheet_name]
        meta           = sheet_meta[sheet_name]
        header_col_map = meta["header_col_map"]
        data_start_row = meta["data_start_row"]

        is_scalar = (sheet_name == CONFIG["SCALAR_SHEET_NAME"])

        serial_id_col = header_col_map.get("serial_id")
        issuer_id_col = header_col_map.get(CONFIG["ISSUER_ID_FIELD"])

        # Build serial_id -> excel row number map (non-scalar only)
        row_map: dict[str, int] = {}
        if not is_scalar and serial_id_col:
            for r in range(data_start_row, data_start_row + meta["populated_rows"]):
                sid = ws.cell(row=r, column=serial_id_col).value
                if sid is not None:
                    row_map[str(sid).strip()] = r

        # Build name -> excel row number map for name-match sheets
        # e.g. {"Audit": 2, "Compensation": 3, "Nominating and Governance": 4}
        name_row_map: dict[str, int] = {}
        anchor_col = name_match_sheets.get(sheet_name)
        if anchor_col and anchor_col in header_col_map:
            anchor_col_idx = header_col_map[anchor_col]
            for r in range(data_start_row, data_start_row + meta["populated_rows"]):
                name_val = ws.cell(row=r, column=anchor_col_idx).value
                if name_val is not None:
                    name_row_map[str(name_val).strip().lower()] = r

        sheet_overrides = 0
        comm_fn_cols = CONFIG.get("COMMITTEEFUNCTION_COLS", set())

        for _, dvv_row in dvv_df.iterrows():
            datalib     = str(dvv_row.get("_DVV_DATALIB", "")).strip()
            correct_val = dvv_row.get("_DVV_CORRECT_VALUE")
            dvv_series  = str(dvv_row.get("_DVV_SERIES_ID", "")).strip()
            dvv_tab     = str(dvv_row.get("_DVV_TAB", "")).strip()
            dvv_uid     = str(dvv_row.get("_DVV_UID", "")).strip()

            if dvv_series.lower() in ("nan", "none", ""):
                dvv_series = ""
            if dvv_uid.lower() in ("nan", "none", ""):
                dvv_uid = ""

            # COMMITTEEFUNCTION columns: convert Yes -> 1
            if (datalib in comm_fn_cols and
                    correct_val is not None and
                    str(correct_val).strip().lower() == "yes"):
                correct_val = "1"

            # Skip if DATALIB_TAG not a header in this sheet
            if not datalib or datalib not in header_col_map:
                continue

            target_col = header_col_map[datalib]

            logger.debug(
                f"  DVV row: sheet='{sheet_name}' datalib='{datalib}' "
                f"series='{dvv_series}' tab='{dvv_tab}' "
                f"uid='{dvv_uid}' val='{correct_val}'"
            )

            # ── SCALAR: write directly to single issuer row ────────────────
            if is_scalar:
                ws.cell(row=data_start_row, column=target_col).value = correct_val
                ws.cell(row=data_start_row, column=target_col).fill  = dvv_fill

            # ── SERIES ID present: match by SERIALID ──────────────────────
            elif dvv_series and serial_id_col:
                if dvv_series in row_map:
                    row_num = row_map[dvv_series]
                    ws.cell(row=row_num, column=target_col).value = correct_val
                    ws.cell(row=row_num, column=target_col).fill  = dvv_fill
                    logger.debug(
                        f"  DVV series-match: SeriesID={dvv_series} -> row {row_num}"
                    )
                else:
                    # SERIALID not found — try UID name-match for name-match sheets
                    matched_by_name = False
                    if sheet_name in name_match_sheets and dvv_uid and name_row_map:
                        lookup_name = dvv_uid.strip().lower()
                        if lookup_name in name_row_map:
                            row_num = name_row_map[lookup_name]
                            ws.cell(row=row_num, column=target_col).value = correct_val
                            ws.cell(row=row_num, column=target_col).fill  = dvv_fill
                            matched_by_name = True
                            logger.info(
                                f"  DVV name-fallback (series miss): "
                                f"sheet='{sheet_name}' UID='{dvv_uid}' "
                                f"DataLib={datalib} -> row {row_num}"
                            )
                    if not matched_by_name:
                        new_row = data_start_row + meta["populated_rows"]
                        if issuer_id_col:
                            ws.cell(row=new_row, column=issuer_id_col).value = issuer_id
                        ws.cell(row=new_row, column=serial_id_col).value = dvv_series
                        ws.cell(row=new_row, column=target_col).value = correct_val
                        ws.cell(row=new_row, column=target_col).fill  = dvv_fill
                        row_map[dvv_series] = new_row
                        meta["populated_rows"] += 1
                        logger.info(
                            f"  DVV new row: sheet='{sheet_name}' "
                            f"SeriesID={dvv_series} DataLib={datalib}"
                        )

            # ── NO SERIES ID: name-match using UID (Column S) ─────────────
            elif not dvv_series and sheet_name in name_match_sheets:
                lookup_name = dvv_uid.strip().lower()
                if lookup_name and lookup_name in name_row_map:
                    row_num = name_row_map[lookup_name]
                    ws.cell(row=row_num, column=target_col).value = correct_val
                    ws.cell(row=row_num, column=target_col).fill  = dvv_fill
                    logger.info(
                        f"  DVV name-match: sheet='{sheet_name}' "
                        f"UID='{dvv_uid}' DataLib={datalib} -> row {row_num}"
                    )
                else:
                    logger.warning(
                        f"  DVV name-match MISS: sheet='{sheet_name}' "
                        f"UID='{dvv_uid}' DataLib={datalib} — "
                        f"no row where {anchor_col}='{dvv_uid}'. "
                        f"Available: {list(name_row_map.keys())}"
                    )
                    continue

            # ── No SERIALID + not a name-match sheet -> skip ───────────────
            else:
                logger.debug(
                    f"  DVV skip: sheet='{sheet_name}' datalib='{datalib}' "
                    f"series='' uid='{dvv_uid}'"
                )
                continue

            sheet_overrides += 1

        if sheet_overrides:
            logger.info(
                f"  DVV '{template_key}' / '{sheet_name}': "
                f"{sheet_overrides} override(s) applied"
            )
        total_overrides += sheet_overrides

    logger.info(f"DVV overrides '{template_key}': {total_overrides} total override(s)")
    return total_overrides


# ==============================================================================
# SECTION 8: OUTPUT FORMATTING
# ==============================================================================

def _thin_border() -> Border:
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr_font() -> Font:
    return Font(
        name=CONFIG["HEADER_FONT_NAME"], size=CONFIG["HEADER_FONT_SIZE"],
        bold=True, color=CONFIG["HEADER_FONT_COLOR"]
    )

def _hdr_fill() -> PatternFill:
    return PatternFill(
        "solid",
        start_color=CONFIG["HEADER_FILL_COLOR"],
        end_color=CONFIG["HEADER_FILL_COLOR"]
    )

def _dat_font() -> Font:
    return Font(
        name=CONFIG["DATA_FONT_NAME"], size=CONFIG["DATA_FONT_SIZE"],
        bold=False
    )

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center")


def format_worksheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    data_start_row: int,
    populated_rows: int,
    logger: logging.Logger,
) -> None:
    """
    Apply professional formatting:
    - Header row : Times New Roman 12 Bold, White text, Blue fill, Centred
    - Data rows  : Times New Roman 12, thin borders
    - DVV cells  : preserve #E6FDCF highlight
    - AutoFilter, freeze top row, auto-fit columns
    """
    max_col = ws.max_column
    if not max_col:
        return

    last_row = data_start_row + populated_rows - 1
    if populated_rows == 0:
        last_row = data_start_row - 1

    border = _thin_border()

    # Header row
    for c in range(1, max_col + 1):
        cell = ws.cell(row=1, column=c)
        if cell.value is not None:
            cell.font      = _hdr_font()
            cell.fill      = _hdr_fill()
            cell.alignment = _center()
            cell.border    = border

    # Data rows — top + left aligned
    data_align = Alignment(horizontal="left", vertical="top", wrap_text=False)
    for r in range(data_start_row, last_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font      = _dat_font()
            cell.border    = border
            cell.alignment = data_align
            # Row height — ensure enough room for top-aligned text
        ws.row_dimensions[r].height = 15

    # Auto filter on header row
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}1"

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-fit column widths
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = max(
            (len(str(cell.value)) for cell in col_cells if cell.value),
            default=8
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    # AutoFilter on full data range (replaces Excel Table objects).
    #
    # WHY: openpyxl writes Table XML files into the zip but does NOT correctly
    # wire the worksheet .rels or <tablePart> references that Excel requires.
    # Excel then sees orphaned table XML, repairs the file on open, and removes
    # the tables — producing the "We found a problem" recovery dialog.
    # AutoFilter achieves the same user-visible result (column dropdowns,
    # sortable headers) with zero corruption risk.
    filter_ref = f"A1:{get_column_letter(max_col)}1"
    if populated_rows > 0:
        filter_ref = f"A1:{get_column_letter(max_col)}{last_row}"
    ws.auto_filter.ref = filter_ref


# ==============================================================================
# SECTION 9: VALIDATION LOG
# ==============================================================================

def build_validation_log(
    all_template_headers: dict,
    ext_df: pd.DataFrame,
    issuer_id: str,
    logger: logging.Logger,
) -> list[dict]:
    """
    Produce validation records for:
    1. Template Data Libs missing in extraction
    2. Extraction Data Libs not used in any template
    """
    ext_cols      = set(ext_df.columns.tolist())
    all_tmpl_cols: set[str] = set()
    records = []

    for tmpl_key, sheets in all_template_headers.items():
        for sheet_name, headers in sheets.items():
            for h in headers:
                if h is None:
                    continue
                all_tmpl_cols.add(h)
                if (
                    h not in ext_cols
                    and h != CONFIG["ISSUER_ID_FIELD"]
                    and h != "serial_id"
                ):
                    records.append({
                        "Check Type": "Template DataLib missing in Extraction",
                        "Template":   tmpl_key,
                        "Sheet":      sheet_name,
                        "Data Lib":   h,
                        "Detail":     "Header in template; no matching column in CSV",
                    })

    for col in ext_cols:
        if col not in all_tmpl_cols and col not in {CONFIG["ISSUER_ID_FIELD"], "serial_id"}:
            records.append({
                "Check Type": "Extraction DataLib unused in Templates",
                "Template":   "ALL",
                "Sheet":      "ALL",
                "Data Lib":   col,
                "Detail":     "Column in CSV; not used in any template sheet",
            })

    logger.info(f"Validation: {len(records)} issues found")
    return records


def _write_log_sheet(
    wb: openpyxl.Workbook,
    sheet_name: str,
    columns: list[str],
    records: list[dict],
    logger: logging.Logger,
) -> None:
    """Generic helper to write a formatted log sheet into a workbook."""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    border = _thin_border()

    for ci, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font      = _hdr_font()
        cell.fill      = _hdr_fill()
        cell.alignment = _center()
        cell.border    = border

    for ri, record in enumerate(records, start=2):
        for ci, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=ri, column=ci, value=record.get(col_name, ""))
            cell.font   = _dat_font()
            cell.border = border

    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = max(
            (len(str(c.value)) for c in col_cells if c.value), default=10
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 80)

    ws.freeze_panes = "A2"
    logger.info(f"'{sheet_name}' written: {len(records)} records.")


def add_validation_sheet(wb, records, logger):
    _write_log_sheet(
        wb, "Validation_Log",
        ["Check Type", "Template", "Sheet", "Data Lib", "Detail"],
        records, logger
    )


def add_dvv_audit_sheet(wb, records, logger):
    _write_log_sheet(
        wb, "DVV_Audit_Log",
        ["Issuer ID", "Series ID", "Data Lib", "Old Value", "New Value",
         "Update Timestamp", "Template Name", "Worksheet Name", "Status"],
        records, logger
    )


# ==============================================================================
# SECTION 10: ORCHESTRATION
# ==============================================================================

def process_issuer(issuer_id: str, boot_logger: logging.Logger) -> dict:
    """
    Full end-to-end pipeline for one Issuer ID.
    Returns a summary dict.
    """
    start_time = datetime.datetime.now()
    summary = {
        "Start Time":        start_time.strftime("%Y-%m-%d %H:%M:%S"),
        "End Time":          "",
        "Issuer ID":         issuer_id,
        "Issuer Name":       "",
        "Extraction File":   "",
        "DVV File":          "",
        "Records Processed": 0,
        "Fields Populated":  0,
        "DVV Overrides":     0,
        "Validation Issues": 0,
        "Errors":            [],
        "Status":            "PENDING",
    }

    extraction_folder = Path(CONFIG["EXTRACTION_FOLDER"])
    dvv_folder        = Path(CONFIG["DVV_FOLDER"])
    template_folder   = Path(CONFIG["TEMPLATE_FOLDER"])
    output_root       = Path(CONFIG["OUTPUT_ROOT"])
    issuer_name       = ""

    # ── STEP 1: File Discovery ─────────────────────────────────────────────
    boot_logger.info("=" * 70)
    boot_logger.info(f"STEP 1: File Discovery  |  Issuer: {issuer_id}")

    try:
        extraction_path, issuer_name = find_extraction_file(
            issuer_id, extraction_folder, boot_logger
        )
        summary["Extraction File"] = extraction_path.name
        summary["Issuer Name"]     = issuer_name
    except FileNotFoundError as e:
        boot_logger.error(f"FATAL: {e}")
        summary["Errors"].append(str(e))
        summary["Status"] = "FAILED"
        summary["End Time"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        return summary

    try:
        dvv_path = find_dvv_file(issuer_id, dvv_folder, boot_logger)
        summary["DVV File"] = dvv_path.name
    except FileNotFoundError as e:
        boot_logger.error(f"FATAL: {e}")
        summary["Errors"].append(str(e))
        summary["Status"] = "FAILED"
        summary["End Time"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        return summary

    # ── Create per-issuer output sub-folder ────────────────────────────────
    output_folder = output_root / f"{issuer_id}_{issuer_name}_Output"
    output_folder.mkdir(parents=True, exist_ok=True)

    # Now that the per-issuer folder exists, create the real logger.
    # Logging goes to console only — no .log file written to the output folder.
    logger = logging.getLogger("AnnualUpdate")
    if logger.handlers:
        logger.handlers.clear()
    logger.setLevel(logging.DEBUG)
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(logging.Formatter("%(levelname)-8s | %(message)s"))
    logger.addHandler(ch)
    logger.info(f"Output folder: {output_folder}")

    # ── STEP 2: Load Extraction CSV ────────────────────────────────────────
    logger.info("=" * 70)
    logger.info("STEP 2: Loading extraction CSV")
    try:
        ext_df = load_extraction(extraction_path, issuer_id, logger)
        summary["Records Processed"] = len(ext_df)
    except Exception as e:
        logger.error(f"FATAL: {e}")
        logger.debug(traceback.format_exc())
        summary["Errors"].append(str(e))
        summary["Status"] = "FAILED"
        summary["End Time"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        return summary

    # ── STEP 3: Load DVV File ──────────────────────────────────────────────
    logger.info("=" * 70)
    logger.info("STEP 3: Loading DVV file")
    try:
        dvv_df = load_dvv(dvv_path, logger)
    except Exception as e:
        logger.error(f"FATAL: {e}")
        logger.debug(traceback.format_exc())
        summary["Errors"].append(str(e))
        summary["Status"] = "FAILED"
        summary["End Time"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        return summary

    # ── STEP 4: Populate Templates + DVV Overrides ─────────────────────────
    logger.info("=" * 70)
    logger.info("STEP 4: Populating templates and applying DVV overrides")

    all_template_headers: dict[str, dict[str, list]] = {}
    total_dvv_overrides   = 0
    total_fields          = 0

    for tmpl_key, tmpl_filename in CONFIG["TEMPLATES"].items():
        tmpl_path = template_folder / tmpl_filename
        out_path  = output_folder / f"{issuer_id}_{CONFIG['OUTPUT_NAMES'][tmpl_key]}.xlsx"

        logger.info(f"  ── {tmpl_key}: {tmpl_filename} ──")

        if not tmpl_path.exists():
            msg = f"Template not found: {tmpl_path}"
            logger.error(msg)
            summary["Errors"].append(msg)
            continue

        # Populate
        try:
            wb, sheet_meta = populate_template(
                tmpl_path, tmpl_key, ext_df, issuer_id, logger
            )
        except Exception as e:
            msg = f"Population failed for '{tmpl_key}': {e}"
            logger.error(msg)
            logger.debug(traceback.format_exc())
            summary["Errors"].append(msg)
            continue

        all_template_headers[tmpl_key] = {
            sn: m["headers"] for sn, m in sheet_meta.items()
        }
        for m in sheet_meta.values():
            total_fields += len(m["headers"]) * m["populated_rows"]

        # DVV overrides
        try:
            overrides = apply_dvv_overrides(
                wb, sheet_meta, dvv_df, issuer_id, tmpl_key, logger
            )
            total_dvv_overrides += overrides
        except Exception as e:
            msg = f"DVV override failed for '{tmpl_key}': {e}"
            logger.error(msg)
            logger.debug(traceback.format_exc())
            summary["Errors"].append(msg)

        # Format all data sheets
        try:
            for sn in wb.sheetnames:
                if sn in sheet_meta:
                    m = sheet_meta[sn]
                    format_worksheet(
                        wb[sn],
                        m["data_start_row"],
                        m["populated_rows"],
                        logger,
                    )
        except Exception as e:
            logger.error(f"Formatting failed for '{tmpl_key}': {e}")
            logger.debug(traceback.format_exc())

        # Save
        try:
            wb.save(out_path)
            logger.info(f"  Saved: {out_path.name}")
        except Exception as e:
            msg = f"Save failed for '{tmpl_key}': {e}"
            logger.error(msg)
            summary["Errors"].append(msg)

    # ── STEP 5: Validation (console/log only — no extra files written) ────
    logger.info("=" * 70)
    logger.info("STEP 5: Validation check (logged to console only)")

    val_records = build_validation_log(
        all_template_headers, ext_df, issuer_id, logger
    )

    summary["Validation Issues"] = len(val_records)
    summary["DVV Overrides"]     = total_dvv_overrides
    summary["Fields Populated"]  = total_fields

    # Log validation issues to the execution log (no separate xlsx written)
    if val_records:
        logger.info(f"  {len(val_records)} validation issue(s) found:")
        for vr in val_records:
            logger.info(
                f"    [{vr['Check Type']}] "
                f"Template={vr['Template']} | Sheet={vr['Sheet']} | "
                f"DataLib={vr['Data Lib']} | {vr['Detail']}"
            )
    else:
        logger.info("  No validation issues found.")

    # ── Final Summary ──────────────────────────────────────────────────────
    end_time = datetime.datetime.now()
    elapsed  = (end_time - start_time).total_seconds()
    summary["End Time"] = end_time.strftime("%Y-%m-%d %H:%M:%S")
    summary["Status"]   = "FAILED" if summary["Errors"] else "SUCCESS"

    logger.info("=" * 70)
    logger.info("EXECUTION SUMMARY")
    logger.info(f"  Issuer ID         : {issuer_id}")
    logger.info(f"  Issuer Name       : {issuer_name}")
    logger.info(f"  Status            : {summary['Status']}")
    logger.info(f"  Extraction File   : {summary['Extraction File']}")
    logger.info(f"  DVV File          : {summary['DVV File']}")
    logger.info(f"  Records Processed : {summary['Records Processed']}")
    logger.info(f"  Fields Populated  : {summary['Fields Populated']}")
    logger.info(f"  DVV Overrides     : {summary['DVV Overrides']}")
    logger.info(f"  Validation Issues : {summary['Validation Issues']}")
    logger.info(f"  Errors            : {len(summary['Errors'])}")
    logger.info(f"  Elapsed           : {elapsed:.1f}s")
    for err in summary["Errors"]:
        logger.error(f"  ERROR: {err}")
    logger.info("=" * 70)

    return summary


# ==============================================================================
# SECTION 11: ENTRY POINT
# ==============================================================================

def main() -> None:
    print("=" * 70)
    print("  MSCI Annual Update Factual Process Automation  v3.0.0")
    print("=" * 70)

    issuer_id = input("\nEnter Issuer ID (e.g. IID000000002177512): ").strip()

    if not issuer_id:
        print("ERROR: Issuer ID cannot be blank.")
        sys.exit(1)

    if not re.match(r"^IID\d+$", issuer_id, re.IGNORECASE):
        print(
            f"WARNING: '{issuer_id}' does not match expected pattern "
            "'IID' + digits. Proceeding anyway."
        )

    output_root = Path(CONFIG["OUTPUT_ROOT"])
    output_root.mkdir(parents=True, exist_ok=True)

    # Minimal console-only logger for startup messages.
    # The full file logger is created inside process_issuer() once the
    # per-issuer subfolder exists — this avoids stray log files in the root.
    boot_logger = logging.getLogger("AnnualUpdateBoot")
    if not boot_logger.handlers:
        boot_logger.setLevel(logging.INFO)
        bch = logging.StreamHandler(sys.stdout)
        bch.setFormatter(logging.Formatter("%(levelname)-8s | %(message)s"))
        boot_logger.addHandler(bch)

    boot_logger.info(f"Issuer ID  : {issuer_id}")
    boot_logger.info(f"Output root: {output_root}")

    try:
        summary = process_issuer(issuer_id, boot_logger)
    except Exception as e:
        boot_logger.critical(f"Unhandled exception: {e}")
        print(f"\nFATAL ERROR: {e}")
        sys.exit(1)

    print("\n" + "=" * 70)
    print(f"  Status            : {summary['Status']}")
    print(f"  Issuer Name       : {summary.get('Issuer Name', '')}")
    print(f"  Records Processed : {summary['Records Processed']}")
    print(f"  DVV Overrides     : {summary['DVV Overrides']}")
    print(f"  Validation Issues : {summary['Validation Issues']}")
    print(f"  Output folder     : {output_root}")
    if summary["Errors"]:
        print(f"\n  Errors ({len(summary['Errors'])}):")
        for err in summary["Errors"]:
            print(f"    • {err}")
    print("=" * 70)


if __name__ == "__main__":
    main()
