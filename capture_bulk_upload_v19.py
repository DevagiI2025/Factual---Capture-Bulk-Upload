"""
================================================================================
Capture Bulk Upload Generator  —  Streamlit Web App
================================================================================
Version  : 19.1  (speed + UI fix)
Engine   : annual_update_processor_v19 logic (in-memory, no local paths)
UI       : pure native Streamlit — zero custom CSS or HTML

Speed improvements over v19.0:
  • DVV file read uses calamine engine (3-5x faster than openpyxl).
    Falls back to openpyxl automatically if calamine is not installed.
  • Auto-fit column widths removed — uses sensible fixed widths instead.
    This alone saves 10-20 seconds on large sheets.

Run:  streamlit run capture_bulk_upload_v19.py
Install calamine for fastest DVV reading:  pip install python-calamine
================================================================================
"""

import io
import re
import traceback
import zipfile

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Check once at import time whether calamine is available
try:
    import python_calamine          # noqa: F401
    _EXCEL_ENGINE = "calamine"
except ImportError:
    _EXCEL_ENGINE = "openpyxl"


# ==============================================================================
# SECTION 1 — CONFIGURATION
# ==============================================================================

CONFIG = {
    "SCALAR_SHEET_NAME":  "Scalar",
    "ISSUER_ID_FIELD":    "DMX_ISSUER_ID",

    # DVV column letters
    "DVV_TAB_COL":        "B",
    "DVV_UID_COL":        "S",
    "DVV_DATALIB_COL":    "G",
    "DVV_CORRECTVAL_COL": "AH",
    "DVV_SERIES_COL":     "AM",

    # DVV rows to exclude
    "DVV_EXCLUDE_TABS": {
        "director attributes",
        "director data - board",
        "positions",
        "individual",
    },

    # Committee name-match: template sheet → anchor column header
    "DVV_NAME_MATCH_SHEETS": {
        "Committee": "COMMITTEENAME",
    },

    # Individual Name insertion
    "INDIVIDUAL_NAME_FIELD":      "REL_INDIVID",
    "INDIVIDUAL_NAME_COL_HEADER": "Individual Name",
    "INDIVIDUAL_NAME_SHEETS": {
        "Committee Membership",
        "Director Attributes",
        "Director Ownership",
        "Compensation",
        "CEO Compensation",
        "CEO Compensation CIC,SEV",
    },

    # COMMITTEEFUNCTION Yes → 1
    "COMMITTEEFUNCTION_COLS": {
        "COMMITTEEFUNCTIONA",
        "COMMITTEEFUNCTIONC",
        "COMMITTEEFUNCTIONE",
        "COMMITTEEFUNCTIONG",
        "COMMITTEEFUNCTIONN",
        "COMMITTEEFUNCTIONRISK",
        "COMMITFUNCTIONHEALTHSAFETY",
        "COMMITTEEFUNCTIONO",
    },

    # Output formatting
    "HEADER_FONT_NAME":    "Times New Roman",
    "HEADER_FONT_SIZE":    12,
    "HEADER_FONT_COLOR":   "FFFFFF",
    "HEADER_FILL_COLOR":   "0070C0",
    "DATA_FONT_NAME":      "Times New Roman",
    "DATA_FONT_SIZE":      12,
    "DVV_HIGHLIGHT_COLOR": "E6FDCF",

    # Fixed column width used instead of auto-fit (much faster)
    "DEFAULT_COL_WIDTH":   20,
}

TEMPLATE_KEYS = ["Position", "Scalar_Series1", "Series2"]

TEMPLATE_LABELS = {
    "Position":       "Position Tab Split",
    "Scalar_Series1": "Scalar & Series 1",
    "Series2":        "Series 2",
}

OUTPUT_SUFFIXES = {
    "Position":       "Position_Tab_Split",
    "Scalar_Series1": "Scalar & Series 1_DP",
    "Series2":        "Series 2- Delete blank cell_DP",
}

OUTPUT_ICONS = {
    "Position":       "📄",
    "Scalar_Series1": "📋",
    "Series2":        "📃",
}


# ==============================================================================
# SECTION 2 — PROCESSING ENGINE
# ==============================================================================

def _col_idx(letter: str) -> int:
    """Excel column letter(s) → 0-based integer index."""
    r = 0
    for ch in letter.upper():
        r = r * 26 + (ord(ch) - ord("A") + 1)
    return r - 1


def standardize_header(raw: str) -> str:
    """
    'Audit Board Member (REL_AUDIT_BOARD)' → 'REL_AUDIT_BOARD'
    'DMX_ISSUER_ID'                        → 'DMX_ISSUER_ID'
    """
    raw = str(raw).strip()
    m = re.search(r"\(([^)]+)\)\s*$", raw)
    return m.group(1).strip() if m else raw


# ── Extraction CSV ────────────────────────────────────────────────────────────

def load_extraction(file_obj) -> tuple:
    try:
        df = pd.read_csv(file_obj, dtype=str, encoding="utf-8-sig")
    except UnicodeDecodeError:
        file_obj.seek(0)
        df = pd.read_csv(file_obj, dtype=str, encoding="latin-1")

    if df.empty:
        raise ValueError("Extraction CSV is empty.")

    df.columns = [standardize_header(c) for c in df.columns]

    iid = ""
    if CONFIG["ISSUER_ID_FIELD"] in df.columns:
        iid = str(df[CONFIG["ISSUER_ID_FIELD"]].iloc[0]).strip()

    return df, iid


# ── DVV Merged XLSX ───────────────────────────────────────────────────────────

def load_dvv(file_obj) -> pd.DataFrame:
    """
    Read DVV file using calamine (fast) if available, else openpyxl.
    Applies two filters:
      1. Keep rows where Correct Value (col AH) is non-blank.
      2. Exclude rows whose TAB (col B) is in DVV_EXCLUDE_TABS.
    """
    try:
        dvv = pd.read_excel(
            file_obj, sheet_name=0, dtype=str,
            header=0, engine=_EXCEL_ENGINE,
        )
    except Exception:
        # calamine may fail on very old .xls files — fall back to openpyxl
        file_obj.seek(0)
        dvv = pd.read_excel(
            file_obj, sheet_name=0, dtype=str,
            header=0, engine="openpyxl",
        )

    if dvv.empty:
        raise ValueError("DVV file is empty.")

    ti  = _col_idx(CONFIG["DVV_TAB_COL"])
    ui  = _col_idx(CONFIG["DVV_UID_COL"])
    di  = _col_idx(CONFIG["DVV_DATALIB_COL"])
    ci  = _col_idx(CONFIG["DVV_CORRECTVAL_COL"])
    si  = _col_idx(CONFIG["DVV_SERIES_COL"])

    needed = max(ti, ui, di, ci, si)
    if dvv.shape[1] <= needed:
        raise ValueError(
            f"DVV file has only {dvv.shape[1]} columns; "
            f"need at least {needed + 1} "
            f"(up to column {CONFIG['DVV_SERIES_COL']})."
        )

    cols = dvv.columns.tolist()
    dvv = dvv.rename(columns={
        cols[ti]: "_DVV_TAB",
        cols[ui]: "_DVV_UID",
        cols[di]: "_DVV_DATALIB",
        cols[ci]: "_DVV_CORRECT_VALUE",
        cols[si]: "_DVV_SERIES_ID",
    })

    # Filter 1: correct value not blank
    dvv = dvv[
        dvv["_DVV_CORRECT_VALUE"].notna() &
        (dvv["_DVV_CORRECT_VALUE"].str.strip() != "")
    ].copy()

    # Filter 2: exclude specific TAB values
    dvv["_tn"] = dvv["_DVV_TAB"].fillna("").str.strip().str.lower()
    dvv = dvv[~dvv["_tn"].isin(CONFIG["DVV_EXCLUDE_TABS"])].drop(columns=["_tn"])

    return dvv


# ── Series ID helpers ─────────────────────────────────────────────────────────

def build_datalib_to_series_map(ext_df: pd.DataFrame) -> dict:
    """Map each datalib column → index of its nearest preceding Series ID column."""
    cols = list(ext_df.columns)
    sid_pos = [
        i for i, c in enumerate(cols)
        if c == "Series ID" or re.match(r"^Series ID(\.\d+)?$", c)
    ]
    result = {}
    for idx, col in enumerate(cols):
        if col == "Series ID" or re.match(r"^Series ID(\.\d+)?$", col):
            continue
        before = [p for p in sid_pos if p < idx]
        if before:
            result[col] = before[-1]
    return result


def build_individual_name_map(ext_df: pd.DataFrame) -> dict:
    """Build { serial_id: director_name } from REL_INDIVID column."""
    individ_col = CONFIG["INDIVIDUAL_NAME_FIELD"]
    if individ_col not in ext_df.columns:
        return {}

    cols = list(ext_df.columns)
    individ_idx = cols.index(individ_col)
    sid_pos = [
        i for i, c in enumerate(cols)
        if c == "Series ID" or re.match(r"^Series ID(\.\d+)?$", c)
    ]
    before = [p for p in sid_pos if p < individ_idx]
    after  = [p for p in sid_pos if p > individ_idx]

    if before:
        sid_col = cols[before[-1]]
    elif after:
        sid_col = cols[after[0]]
    else:
        return {}

    name_map = {}
    for _, row in ext_df.iterrows():
        sid  = row.get(sid_col)
        name = row.get(individ_col)
        if (pd.notna(sid) and pd.notna(name) and
                str(sid).strip() not in ("", "nan") and
                str(name).strip() not in ("", "nan")):
            name_map[str(sid).strip()] = str(name).strip()
    return name_map


def row_has_data(ext_row, headers: list) -> bool:
    skip = {CONFIG["ISSUER_ID_FIELD"], "serial_id"}
    for h in headers:
        if h in skip:
            continue
        v = ext_row.get(h)
        if pd.notna(v) and str(v).strip() not in ("", "nan"):
            return True
    return False


# ── Template population ───────────────────────────────────────────────────────

def populate_template(tmpl_bytes: bytes, tmpl_key: str,
                      ext_df: pd.DataFrame, issuer_id: str):
    wb = load_workbook(io.BytesIO(tmpl_bytes))
    sheet_meta = {}

    dl_to_sid   = build_datalib_to_series_map(ext_df)
    name_map    = build_individual_name_map(ext_df)
    comm_fn     = CONFIG["COMMITTEEFUNCTION_COLS"]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        headers = [c.value for c in ws[1] if c.value is not None]
        if not headers:
            continue

        hcm = {c.value: i for i, c in enumerate(ws[1], 1) if c.value}
        is_scalar    = (sheet_name == CONFIG["SCALAR_SHEET_NAME"])
        indiv_col_idx = None

        # Insert Individual Name column after serial_id where configured
        if (not is_scalar
                and "serial_id" in hcm
                and sheet_name in CONFIG["INDIVIDUAL_NAME_SHEETS"]):
            ih = CONFIG["INDIVIDUAL_NAME_COL_HEADER"]
            if ih not in hcm:
                ins = hcm["serial_id"] + 1
                ws.insert_cols(ins)
                ws.cell(row=1, column=ins).value = ih
                hcm = {h: (ci + 1 if ci >= ins else ci) for h, ci in hcm.items()}
                hcm[ih] = ins
                indiv_col_idx = ins
            else:
                indiv_col_idx = hcm[ih]

        headers = [c.value for c in ws[1] if c.value is not None]
        data_start_row = 2
        populated_rows = 0

        # ── Scalar sheet: one issuer row ──────────────────────────────────────
        if is_scalar:
            rows = (ext_df[ext_df[CONFIG["ISSUER_ID_FIELD"]] == issuer_id]
                    if CONFIG["ISSUER_ID_FIELD"] in ext_df.columns else ext_df)
            if rows.empty:
                rows = ext_df
            if rows.empty:
                continue
            er = rows.iloc[0]
            for th, ci in hcm.items():
                if th == CONFIG["ISSUER_ID_FIELD"]:
                    ws.cell(row=data_start_row, column=ci).value = issuer_id
                elif th in ext_df.columns:
                    v = er.get(th)
                    ws.cell(row=data_start_row, column=ci).value = (
                        None if (pd.isna(v) or str(v).strip() in ("", "nan"))
                        else str(v).strip()
                    )
            populated_rows = 1

        # ── Series / Position: one row per extraction row ─────────────────────
        else:
            all_cols     = list(ext_df.columns)
            sid_positions = [
                i for i, c in enumerate(all_cols)
                if c == "Series ID" or re.match(r"^Series ID(\.\d+)?$", c)
            ]

            # Best-match section detection
            sheet_sid_col = None
            sheet_sec_cols = []
            best = 0
            for sp in sid_positions:
                nxt = next((p for p in sid_positions if p > sp), len(all_cols))
                cands = [all_cols[i] for i in range(sp + 1, nxt) if all_cols[i] in headers]
                if len(cands) > best:
                    best = len(cands)
                    sheet_sid_col  = sp
                    sheet_sec_cols = cands

            for _, er in ext_df.iterrows():
                # Gate 1: section Series ID non-blank
                if sheet_sid_col is not None:
                    sv = er.get(ext_df.columns[sheet_sid_col])
                    if pd.isna(sv) or str(sv).strip() in ("", "nan"):
                        continue
                    sid_val = str(sv).strip()
                else:
                    sid_val = None

                # Gate 2: at least one data column non-blank
                if sheet_sec_cols:
                    if not any(
                        pd.notna(er.get(c)) and
                        str(er.get(c, "")).strip() not in ("", "nan")
                        for c in sheet_sec_cols
                    ):
                        continue
                elif not row_has_data(er, headers):
                    continue

                wr = data_start_row + populated_rows
                for th, ci in hcm.items():
                    if th == CONFIG["ISSUER_ID_FIELD"]:
                        ws.cell(row=wr, column=ci).value = issuer_id

                    elif th == "serial_id":
                        ws.cell(row=wr, column=ci).value = sid_val
                        if indiv_col_idx and sid_val:
                            nm = name_map.get(sid_val, "")
                            ws.cell(row=wr, column=indiv_col_idx).value = nm or None

                    elif th == CONFIG["INDIVIDUAL_NAME_COL_HEADER"]:
                        pass  # written above with serial_id

                    elif th in ext_df.columns:
                        v = er.get(th)
                        if pd.isna(v) or str(v).strip() in ("", "nan"):
                            ws.cell(row=wr, column=ci).value = None
                        else:
                            clean = str(v).strip()
                            if th in comm_fn and clean.lower() == "yes":
                                clean = "1"
                            ws.cell(row=wr, column=ci).value = clean

                populated_rows += 1

        sheet_meta[sheet_name] = {
            "headers":        headers,
            "header_col_map": hcm,
            "data_start_row": data_start_row,
            "populated_rows": populated_rows,
        }

    return wb, sheet_meta


# ── DVV overrides ─────────────────────────────────────────────────────────────

def apply_dvv_overrides(wb, sheet_meta: dict,
                        dvv_df: pd.DataFrame, issuer_id: str) -> int:
    dvv_fill   = PatternFill("solid",
                             start_color=CONFIG["DVV_HIGHLIGHT_COLOR"],
                             end_color=CONFIG["DVV_HIGHLIGHT_COLOR"])
    nm_sheets  = CONFIG["DVV_NAME_MATCH_SHEETS"]
    comm_fn    = CONFIG["COMMITTEEFUNCTION_COLS"]
    total      = 0

    for sheet_name in wb.sheetnames:
        if sheet_name not in sheet_meta:
            continue

        ws  = wb[sheet_name]
        meta = sheet_meta[sheet_name]
        hcm  = meta["header_col_map"]
        dsr  = meta["data_start_row"]
        is_scalar  = (sheet_name == CONFIG["SCALAR_SHEET_NAME"])
        sid_col    = hcm.get("serial_id")
        iid_col    = hcm.get(CONFIG["ISSUER_ID_FIELD"])

        # serial_id → row map
        row_map: dict = {}
        if not is_scalar and sid_col:
            for r in range(dsr, dsr + meta["populated_rows"]):
                sv = ws.cell(row=r, column=sid_col).value
                if sv is not None:
                    row_map[str(sv).strip()] = r

        # name → row map (committee sheets)
        name_row_map: dict = {}
        anchor = nm_sheets.get(sheet_name)
        if anchor and anchor in hcm:
            for r in range(dsr, dsr + meta["populated_rows"]):
                nv = ws.cell(row=r, column=hcm[anchor]).value
                if nv is not None:
                    name_row_map[str(nv).strip().lower()] = r

        sheet_ov = 0

        for _, dr in dvv_df.iterrows():
            datalib  = str(dr.get("_DVV_DATALIB",      "")).strip()
            corr_val = dr.get("_DVV_CORRECT_VALUE")
            dvv_sid  = str(dr.get("_DVV_SERIES_ID",    "")).strip()
            dvv_uid  = str(dr.get("_DVV_UID",          "")).strip()

            if dvv_sid.lower() in ("nan", "none", ""): dvv_sid = ""
            if dvv_uid.lower() in ("nan", "none", ""): dvv_uid = ""

            # COMMITTEEFUNCTION Yes → 1
            if (datalib in comm_fn and corr_val is not None
                    and str(corr_val).strip().lower() == "yes"):
                corr_val = "1"

            if not datalib or datalib not in hcm:
                continue

            tc = hcm[datalib]

            if is_scalar:
                ws.cell(row=dsr, column=tc).value = corr_val
                ws.cell(row=dsr, column=tc).fill  = dvv_fill
                sheet_ov += 1

            elif dvv_sid and sid_col:
                if dvv_sid in row_map:
                    r = row_map[dvv_sid]
                    ws.cell(row=r, column=tc).value = corr_val
                    ws.cell(row=r, column=tc).fill  = dvv_fill
                    sheet_ov += 1
                else:
                    matched = False
                    if sheet_name in nm_sheets and dvv_uid:
                        lk = dvv_uid.lower()
                        if lk in name_row_map:
                            r = name_row_map[lk]
                            ws.cell(row=r, column=tc).value = corr_val
                            ws.cell(row=r, column=tc).fill  = dvv_fill
                            matched = True
                            sheet_ov += 1
                    if not matched:
                        nr = dsr + meta["populated_rows"]
                        if iid_col:
                            ws.cell(row=nr, column=iid_col).value = issuer_id
                        ws.cell(row=nr, column=sid_col).value = dvv_sid
                        ws.cell(row=nr, column=tc).value      = corr_val
                        ws.cell(row=nr, column=tc).fill       = dvv_fill
                        row_map[dvv_sid] = nr
                        meta["populated_rows"] += 1
                        sheet_ov += 1

            elif not dvv_sid and sheet_name in nm_sheets:
                lk = dvv_uid.lower()
                if lk and lk in name_row_map:
                    r = name_row_map[lk]
                    ws.cell(row=r, column=tc).value = corr_val
                    ws.cell(row=r, column=tc).fill  = dvv_fill
                    sheet_ov += 1

        total += sheet_ov

    return total


# ── Formatting (fixed widths — much faster than auto-fit) ────────────────────

def _thin_border() -> Border:
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr_font() -> Font:
    return Font(name=CONFIG["HEADER_FONT_NAME"], size=CONFIG["HEADER_FONT_SIZE"],
                bold=True, color=CONFIG["HEADER_FONT_COLOR"])

def _hdr_fill() -> PatternFill:
    return PatternFill("solid", start_color=CONFIG["HEADER_FILL_COLOR"],
                       end_color=CONFIG["HEADER_FILL_COLOR"])

def _dat_font() -> Font:
    return Font(name=CONFIG["DATA_FONT_NAME"], size=CONFIG["DATA_FONT_SIZE"])


def format_worksheet(ws, data_start_row: int, populated_rows: int) -> None:
    max_col = ws.max_column
    if not max_col:
        return

    last_row   = (data_start_row + populated_rows - 1) if populated_rows > 0 \
                 else (data_start_row - 1)
    border     = _thin_border()
    data_align = Alignment(horizontal="left", vertical="top", wrap_text=False)

    # Header row
    for c in range(1, max_col + 1):
        cell = ws.cell(row=1, column=c)
        if cell.value is not None:
            cell.font      = _hdr_font()
            cell.fill      = _hdr_fill()
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = border

    # Data rows
    for r in range(data_start_row, last_row + 1):
        ws.row_dimensions[r].height = 15
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font      = _dat_font()
            cell.border    = border
            cell.alignment = data_align

    # Fixed column widths (replaces slow auto-fit loop)
    w = CONFIG["DEFAULT_COL_WIDTH"]
    for c in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # AutoFilter on full data range
    end = f"A1:{get_column_letter(max_col)}{last_row}" if populated_rows > 0 \
          else f"A1:{get_column_letter(max_col)}1"
    ws.auto_filter.ref = end
    ws.freeze_panes    = "A2"


# ── Validation log ────────────────────────────────────────────────────────────

def build_validation_log(all_template_headers: dict,
                         ext_df: pd.DataFrame) -> list:
    ext_cols      = set(ext_df.columns)
    all_tmpl_cols: set = set()
    records = []

    for tmpl_key, sheets in all_template_headers.items():
        for sheet_name, headers in sheets.items():
            for h in headers:
                if h is None:
                    continue
                all_tmpl_cols.add(h)
                if (h not in ext_cols
                        and h != CONFIG["ISSUER_ID_FIELD"]
                        and h != "serial_id"
                        and h != CONFIG["INDIVIDUAL_NAME_COL_HEADER"]):
                    records.append({
                        "Check":    "Template DataLib missing in Extraction",
                        "Template": tmpl_key,
                        "Sheet":    sheet_name,
                        "DataLib":  h,
                        "Detail":   "Header in template; no matching column in CSV",
                    })

    skip = {CONFIG["ISSUER_ID_FIELD"], "serial_id",
            CONFIG["INDIVIDUAL_NAME_COL_HEADER"]}
    for col in ext_cols:
        if (col not in all_tmpl_cols and col not in skip
                and not re.match(r"^Series ID(\.\d+)?$", col)):
            records.append({
                "Check":    "Extraction DataLib unused in Templates",
                "Template": "ALL",
                "Sheet":    "ALL",
                "DataLib":  col,
                "Detail":   "Column in CSV; not used in any template",
            })

    return records


# ── Pipeline orchestrator ─────────────────────────────────────────────────────

def run_pipeline(ext_bytes: bytes, dvv_bytes: bytes,
                 template_files: dict) -> dict:
    result = {
        "issuer_id":  "",
        "outputs":    {},
        "stats":      {},
        "errors":     [],
        "validation": [],
    }

    try:
        ext_df, iid = load_extraction(io.BytesIO(ext_bytes))
        result["issuer_id"] = iid
    except Exception as e:
        result["errors"].append(f"Extraction CSV error: {e}")
        return result

    try:
        dvv_df = load_dvv(io.BytesIO(dvv_bytes))
    except Exception as e:
        result["errors"].append(f"DVV file error: {e}")
        return result

    total_ov, total_rows    = 0, 0
    all_template_headers    = {}

    for key in TEMPLATE_KEYS:
        tb = template_files.get(key)
        if not tb:
            result["errors"].append(f"Template missing: {TEMPLATE_LABELS[key]}")
            continue

        try:
            wb, sm = populate_template(tb, key, ext_df, iid)
        except Exception as e:
            result["errors"].append(
                f"{TEMPLATE_LABELS[key]} — population failed: {e}"
            )
            continue

        all_template_headers[key] = {sn: m["headers"] for sn, m in sm.items()}

        try:
            total_ov   += apply_dvv_overrides(wb, sm, dvv_df, iid)
            total_rows += sum(m["populated_rows"] for m in sm.values())
        except Exception as e:
            result["errors"].append(f"{TEMPLATE_LABELS[key]} — DVV override error: {e}")

        try:
            for sn, meta in sm.items():
                format_worksheet(wb[sn], meta["data_start_row"],
                                 meta["populated_rows"])
        except Exception as e:
            result["errors"].append(f"{TEMPLATE_LABELS[key]} — formatting error: {e}")

        try:
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            fname = f"{iid}_{OUTPUT_SUFFIXES[key]}.xlsx"
            result["outputs"][key] = (fname, buf.read())
        except Exception as e:
            result["errors"].append(f"{TEMPLATE_LABELS[key]} — save error: {e}")

    result["stats"]      = {"records": total_rows, "overrides": total_ov}
    result["validation"] = build_validation_log(all_template_headers, ext_df)
    return result


# ==============================================================================
# SECTION 3 — UI  (pure native Streamlit — zero CSS or HTML)
# ==============================================================================

def main():
    st.set_page_config(page_title="Capture Bulk Upload Generator", page_icon="📋")

    if "result" not in st.session_state:
        st.session_state.result = None

    # ── Header ────────────────────────────────────────────────────────────────
    st.title("📋 Capture Bulk Upload Generator")
    st.caption(
        f"Annual Update · Factual Process Automation · v19  "
        f"(Excel engine: **{_EXCEL_ENGINE}**)"
    )
    st.divider()

    # ── Step 1: Source files ──────────────────────────────────────────────────
    st.subheader("Step 1 — Source Files")

    col1, col2 = st.columns(2)
    with col1:
        ext_file = st.file_uploader("Extraction CSV", type=["csv"])
    with col2:
        dvv_file = st.file_uploader("DVV Merged XLSX", type=["xlsx"])

    src_done = ext_file is not None and dvv_file is not None
    st.divider()

    # ── Step 2: Templates ─────────────────────────────────────────────────────
    st.subheader("Step 2 — Bulk Upload Templates")

    pos_file = sc_file = s2_file = None

    if not src_done:
        st.info("Complete Step 1 first.")
    else:
        col3, col4, col5 = st.columns(3)
        with col3:
            pos_file = st.file_uploader("Position Tab Split", type=["xlsx"])
        with col4:
            sc_file  = st.file_uploader("Scalar & Series 1",  type=["xlsx"])
        with col5:
            s2_file  = st.file_uploader("Series 2",           type=["xlsx"])

    all_ready  = src_done and all(f is not None for f in [pos_file, sc_file, s2_file])
    files_done = sum(1 for f in [ext_file, dvv_file, pos_file, sc_file, s2_file]
                     if f is not None)
    st.progress(files_done / 5, text=f"{files_done} of 5 files uploaded")
    st.divider()

    # ── Step 3: Generate ──────────────────────────────────────────────────────
    st.subheader("Step 3 — Generate & Download")

    if not all_ready:
        st.warning(f"Upload {5 - files_done} more file(s) above to continue.")
    else:
        if st.button("⚡ Generate Output Files", type="primary"):
            with st.spinner("Processing — please wait…"):
                result = run_pipeline(
                    ext_bytes      = ext_file.read(),
                    dvv_bytes      = dvv_file.read(),
                    template_files = {
                        "Position":       pos_file.read(),
                        "Scalar_Series1": sc_file.read(),
                        "Series2":        s2_file.read(),
                    },
                )
            st.session_state.result = result
            st.rerun()

    # ── Results ───────────────────────────────────────────────────────────────
    result = st.session_state.result
    if result is None:
        return

    for e in result.get("errors", []):
        st.error(e)

    outputs = result.get("outputs", {})
    if not outputs:
        return

    iid   = result.get("issuer_id", "")
    stats = result.get("stats", {})

    st.success(
        f"Files ready!  Issuer: **{iid}** · "
        f"{stats.get('records', 0)} rows · "
        f"{stats.get('overrides', 0)} DVV overrides applied"
    )

    mc1, mc2, mc3 = st.columns(3)
    mc1.metric("Output Files",  len(outputs))
    mc2.metric("Rows Written",  stats.get("records",   0))
    mc3.metric("DVV Overrides", stats.get("overrides", 0))

    st.write("")
    st.write("**Download individual files:**")
    for key in TEMPLATE_KEYS:
        if key in outputs:
            fname, fbytes = outputs[key]
            st.download_button(
                label     = f"{OUTPUT_ICONS[key]}  {TEMPLATE_LABELS[key]}",
                data      = fbytes,
                file_name = fname,
                mime      = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key       = f"dl_{key}",
            )

    st.write("**Or download all 3 together:**")
    z = io.BytesIO()
    with zipfile.ZipFile(z, "w", zipfile.ZIP_DEFLATED) as zf:
        for key in TEMPLATE_KEYS:
            if key in outputs:
                fname, fbytes = outputs[key]
                zf.writestr(fname, fbytes)
    z.seek(0)
    st.download_button(
        label     = "⬇ Download all 3 as ZIP",
        data      = z.read(),
        file_name = f"{iid}_BulkUpload.zip",
        mime      = "application/zip",
        key       = "dl_zip",
    )

    validation = result.get("validation", [])
    if validation:
        with st.expander(f"Validation notes ({len(validation)})"):
            st.dataframe(
                pd.DataFrame(validation)[
                    ["Check", "Template", "Sheet", "DataLib", "Detail"]
                ],
                use_container_width=True,
                hide_index=True,
            )

    st.divider()
    st.caption("DVV-overridden cells are highlighted green in all output files.")


# ==============================================================================
# ENTRY POINT
# ==============================================================================

if __name__ == "__main__":
    import sys
    import os
    script = os.path.abspath(__file__)
    quoted = f'"{script}"' if " " in script else script
    print(f"\n  Run with:  streamlit run {quoted}\n")
    sys.exit(0)
